'''
Created on Apr 12, 2020

@author: Nitesh
'''

import openpyxl
import traceback
import os
from PyTestFramework.crmpro_automation.filereader.searchfile import get_file
from PyTestFramework.crmpro_automation.utilities.general_custom_exceptions import *

class ExcelReader():
    __data_dict={}
    
    def __init__(self, file_name, sheet_name, testname):
        self.read_file_data(file_name, sheet_name, testname)
        
    def get_test_data(self):
        return self.__data_dict
    
    def read_file_data(self, file_name, sheet_name, testname):
        cur_dir_path = os.path.dirname(os.path.realpath(__file__))
        dir_path=cur_dir_path.split(sep="\\filereader")[0]+"\\testdata"
        excel_file_path = str(get_file(dir_path, file_name)).replace("\\", "\\\\")
        wb = openpyxl.load_workbook(excel_file_path)
        self.sheet = wb[sheet_name]
        self.__data_dict=self.read_test_data(testname)

    def read_test_data(self, testname):
        data = []
        try:
            test_row_num = []
            for row_num, row in enumerate(self.sheet):
                cell = row[0]
                if cell.value == testname:
                    test_row_num.append(row_num+1)

            for testRowNum in test_row_num:
                dict = {}
                for colNum in range(1, self.sheet.max_column+1):
                    cell = self.sheet.cell(row=testRowNum, column=colNum)
                    column_name = self.sheet.cell(row=1, column=colNum).value
                    cell_value = cell.value
                    dict[column_name] = cell_value
                data.append(dict)
        except:
            print(traceback.print_exc())

        return data